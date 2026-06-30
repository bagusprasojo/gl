from io import BytesIO

from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def export_rows_xlsx(filename, title, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    ws.append([title])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_rows_pdf(filename, title, headers, rows):
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    y = height - 40
    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(40, y, title)
    y -= 28
    pdf.setFont('Helvetica-Bold', 8)
    pdf.drawString(40, y, ' | '.join(headers))
    y -= 18
    pdf.setFont('Helvetica', 8)
    for row in rows:
        if y < 40:
            pdf.showPage()
            y = height - 40
            pdf.setFont('Helvetica', 8)
        text = ' | '.join(str(value) for value in row)
        pdf.drawString(40, y, text[:130])
        y -= 14
    pdf.save()
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def format_number_id(value):
    if value in [None, '']:
        return ''
    try:
        amount = round(value)
    except TypeError:
        return value
    return f'{amount:,}'.replace(',', '.')


def financial_statement_xlsx(filename, company, title, period_label, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin_gray = Side(style='thin', color='CBD5E1')
    border = Border(bottom=thin_gray)
    header_fill = PatternFill('solid', fgColor='E2E8F0')

    ws['A1'] = company.name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = company.legal_name or ''
    ws['A3'] = company.address or ''
    ws['A4'] = f'NPWP: {company.tax_number}' if company.tax_number else ''
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=13)
    ws['A7'] = period_label

    ws['A9'] = 'Item'
    ws['B9'] = 'Nilai'
    for cell in ws[9]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border

    row_number = 10
    for row in rows:
        if row.get('is_blank'):
            row_number += 1
            continue
        label_cell = ws.cell(row=row_number, column=1, value=row['label'])
        value_cell = ws.cell(row=row_number, column=2, value=row.get('value') if row.get('has_value') else None)
        label_cell.alignment = Alignment(indent=row.get('indent_level', 0))
        value_cell.alignment = Alignment(horizontal='right')
        value_cell.number_format = '#,##0'
        if row.get('is_bold'):
            label_cell.font = Font(bold=True)
            value_cell.font = Font(bold=True)
        label_cell.border = border
        value_cell.border = border
        row_number += 1

    ws.column_dimensions['A'].width = 48
    ws.column_dimensions['B'].width = 18

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def financial_statement_pdf(filename, company, title, period_label, rows):
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        return HttpResponse(
            'PDF export requires WeasyPrint native dependencies on Windows '
            '(GTK/Pango, including libgobject-2.0-0). Excel export is available.',
            status=503,
            content_type='text/plain',
        )

    html = render_to_string(
        'accounting/exports/financial_statement_pdf.html',
        {
            'company': company,
            'title': title,
            'period_label': period_label,
            'rows': [
                {
                    **row,
                    'formatted_value': format_number_id(row.get('value')) if row.get('has_value') else '',
                }
                for row in rows
            ],
        },
    )
    output = BytesIO()
    HTML(string=html).write_pdf(output)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def export_financial_statement(request, filename, company, title, period_label, rows):
    export_format = request.GET.get('export')
    if export_format == 'xlsx':
        return financial_statement_xlsx(filename, company, title, period_label, rows)
    if export_format == 'pdf':
        return financial_statement_pdf(filename, company, title, period_label, rows)
    return None


def trial_balance_xlsx(filename, company, title, period_label, rows, total_debit, total_credit, is_balanced, difference):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin_gray = Side(style='thin', color='CBD5E1')
    border = Border(bottom=thin_gray)
    header_fill = PatternFill('solid', fgColor='E2E8F0')

    ws['A1'] = company.name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = company.legal_name or ''
    ws['A3'] = company.address or ''
    ws['A4'] = f'NPWP: {company.tax_number}' if company.tax_number else ''
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=13)
    ws['A7'] = period_label
    ws['A8'] = 'Status'
    ws['B8'] = 'Seimbang' if is_balanced else f'Tidak seimbang, selisih {format_number_id(difference)}'

    headers = ['Kode', 'Akun', 'Debit', 'Kredit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border

    row_number = 11
    for row in rows:
        ws.cell(row=row_number, column=1, value=row['account'].code).border = border
        ws.cell(row=row_number, column=2, value=row['account'].name).border = border
        debit_cell = ws.cell(row=row_number, column=3, value=row['debit'])
        credit_cell = ws.cell(row=row_number, column=4, value=row['credit'])
        for cell in [debit_cell, credit_cell]:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0'
            cell.border = border
        row_number += 1

    ws.cell(row=row_number, column=1, value='Total')
    ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
    ws.cell(row=row_number, column=3, value=total_debit)
    ws.cell(row=row_number, column=4, value=total_credit)
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.border = border
        if cell.column >= 3:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0'

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def trial_balance_pdf(filename, company, title, period_label, rows, total_debit, total_credit, is_balanced, difference):
    output = BytesIO()
    status = 'Seimbang' if is_balanced else f'Tidak seimbang, selisih {format_number_id(difference)}'
    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TrialTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=13, leading=16, alignment=1, spaceAfter=2))
    styles.add(ParagraphStyle(name='TrialCompany', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=12, alignment=1))
    styles.add(ParagraphStyle(name='TrialInfo', parent=styles['Normal'], fontSize=8, leading=10, alignment=1))
    styles.add(ParagraphStyle(name='TrialSmall', parent=styles['Normal'], fontSize=8, leading=10))

    story = [Paragraph(company.name, styles['TrialCompany'])]
    if company.legal_name:
        story.append(Paragraph(company.legal_name, styles['TrialInfo']))
    if company.address:
        story.append(Paragraph(company.address, styles['TrialInfo']))
    if company.tax_number:
        story.append(Paragraph(f'NPWP: {company.tax_number}', styles['TrialInfo']))
    story.extend([
        Spacer(1, 6 * mm),
        Paragraph(title.upper(), styles['TrialTitle']),
        Paragraph(period_label, styles['TrialInfo']),
        Paragraph(f'Status: {status}', styles['TrialInfo']),
        Spacer(1, 5 * mm),
    ])

    table_data = [['Kode', 'Akun', 'Debit', 'Kredit']]
    for row in rows:
        table_data.append([
            row['account'].code,
            Paragraph(row['account'].name, styles['TrialSmall']),
            format_number_id(row['debit']),
            format_number_id(row['credit']),
        ])
    table_data.append(['Total', '', format_number_id(total_debit), format_number_id(total_credit)])

    last_row = len(table_data) - 1
    table = Table(
        table_data,
        colWidths=[25 * mm, 82 * mm, 34 * mm, 34 * mm],
        repeatRows=1,
    )
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#94A3B8')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, last_row), (1, last_row)),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(table)

    def add_page_footer(pdf_canvas, doc):
        pdf_canvas.saveState()
        pdf_canvas.setFont('Helvetica', 7)
        pdf_canvas.drawRightString(A4[0] - 15 * mm, 8 * mm, f'Halaman {doc.page}')
        pdf_canvas.restoreState()

    document.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def export_trial_balance(request, filename, company, title, period_label, data):
    export_format = request.GET.get('export')
    if export_format == 'xlsx':
        return trial_balance_xlsx(filename, company, title, period_label, **data)
    if export_format == 'pdf':
        return trial_balance_pdf(filename, company, title, period_label, **data)
    return None


def six_column_trial_balance_xlsx(filename, company, title, period_label, rows, totals, opening_is_balanced, movement_is_balanced, closing_is_balanced):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    thin_gray = Side(style='thin', color='CBD5E1')
    border = Border(bottom=thin_gray)
    header_fill = PatternFill('solid', fgColor='E2E8F0')

    ws['A1'] = company.name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = company.legal_name or ''
    ws['A3'] = company.address or ''
    ws['A4'] = f'NPWP: {company.tax_number}' if company.tax_number else ''
    ws['A6'] = title
    ws['A6'].font = Font(bold=True, size=13)
    ws['A7'] = period_label
    ws['A8'] = 'Status'
    ws['B8'] = 'Seimbang' if closing_is_balanced else 'Tidak seimbang'

    headers = ['Kode', 'Akun', 'Awal Debit', 'Awal Kredit', 'Mutasi Debit', 'Mutasi Kredit', 'Akhir Debit', 'Akhir Kredit']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border

    row_number = 11
    for row in rows:
        values = [
            row['account'].code,
            row['account'].name,
            row['opening_debit'],
            row['opening_credit'],
            row['movement_debit'],
            row['movement_credit'],
            row['closing_debit'],
            row['closing_credit'],
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_number, column=col, value=value)
            cell.border = border
            if col >= 3:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0'
        row_number += 1

    ws.cell(row=row_number, column=1, value='Total')
    ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
    for col, key in enumerate(totals, start=3):
        ws.cell(row=row_number, column=col, value=totals[key])
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.border = border
        if cell.column >= 3:
            cell.alignment = Alignment(horizontal='right')
            cell.number_format = '#,##0'

    widths = [14, 38, 16, 16, 16, 16, 16, 16]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def six_column_trial_balance_pdf(filename, company, title, period_label, rows, totals, opening_is_balanced, movement_is_balanced, closing_is_balanced):
    output = BytesIO()
    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='ReportTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=13, leading=16, alignment=1, spaceAfter=2))
    styles.add(ParagraphStyle(name='CompanyName', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=12, alignment=1))
    styles.add(ParagraphStyle(name='CompanyInfo', parent=styles['Normal'], fontSize=8, leading=10, alignment=1))
    styles.add(ParagraphStyle(name='SmallText', parent=styles['Normal'], fontSize=7, leading=9))
    styles.add(ParagraphStyle(name='SmallRight', parent=styles['SmallText'], alignment=2))

    story = [
        Paragraph(company.name, styles['CompanyName']),
    ]
    if company.legal_name:
        story.append(Paragraph(company.legal_name, styles['CompanyInfo']))
    if company.address:
        story.append(Paragraph(company.address, styles['CompanyInfo']))
    if company.tax_number:
        story.append(Paragraph(f'NPWP: {company.tax_number}', styles['CompanyInfo']))
    story.extend([
        Spacer(1, 5 * mm),
        Paragraph(title.upper(), styles['ReportTitle']),
        Paragraph(period_label, styles['CompanyInfo']),
        Paragraph(f"Status saldo akhir: {'Seimbang' if closing_is_balanced else 'Tidak seimbang'}", styles['CompanyInfo']),
        Spacer(1, 5 * mm),
    ])

    table_data = [
        ['Kode', 'Akun', 'Saldo Awal', '', 'Mutasi', '', 'Saldo Akhir', ''],
        ['', '', 'Debit', 'Kredit', 'Debit', 'Kredit', 'Debit', 'Kredit'],
    ]
    for row in rows:
        table_data.append([
            row['account'].code,
            Paragraph(row['account'].name, styles['SmallText']),
            format_number_id(row['opening_debit']),
            format_number_id(row['opening_credit']),
            format_number_id(row['movement_debit']),
            format_number_id(row['movement_credit']),
            format_number_id(row['closing_debit']),
            format_number_id(row['closing_credit']),
        ])
    table_data.append([
        'Total',
        '',
        format_number_id(totals['opening_debit']),
        format_number_id(totals['opening_credit']),
        format_number_id(totals['movement_debit']),
        format_number_id(totals['movement_credit']),
        format_number_id(totals['closing_debit']),
        format_number_id(totals['closing_credit']),
    ])

    last_row = len(table_data) - 1
    table = Table(
        table_data,
        colWidths=[18 * mm, 49 * mm, 28 * mm, 28 * mm, 28 * mm, 28 * mm, 28 * mm, 28 * mm],
        repeatRows=2,
    )
    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#94A3B8')),
        ('BACKGROUND', (0, 0), (-1, 1), colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, last_row), (-1, last_row), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTNAME', (0, last_row), (-1, last_row), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('ALIGN', (2, 2), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('SPAN', (0, 0), (0, 1)),
        ('SPAN', (1, 0), (1, 1)),
        ('SPAN', (2, 0), (3, 0)),
        ('SPAN', (4, 0), (5, 0)),
        ('SPAN', (6, 0), (7, 0)),
        ('SPAN', (0, last_row), (1, last_row)),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(table)

    def add_page_footer(pdf_canvas, doc):
        pdf_canvas.saveState()
        pdf_canvas.setFont('Helvetica', 7)
        pdf_canvas.drawRightString(landscape(A4)[0] - 12 * mm, 7 * mm, f'Halaman {doc.page}')
        pdf_canvas.restoreState()

    document.build(story, onFirstPage=add_page_footer, onLaterPages=add_page_footer)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return response


def export_six_column_trial_balance(request, filename, company, title, period_label, data):
    export_format = request.GET.get('export')
    if export_format == 'xlsx':
        return six_column_trial_balance_xlsx(filename, company, title, period_label, **data)
    if export_format == 'pdf':
        return six_column_trial_balance_pdf(filename, company, title, period_label, **data)
    return None
